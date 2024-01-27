import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = 1
    while True:
        if sheet.cell(rows, 1).value:
            rows += 1
        else:
            rows -= 1
            break
    return rows

def getColsCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    cols = 1
    while True:
        if sheet.cell(1, cols).value:
            cols += 1
        else:
            cols -= 1
            break
    return cols

def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columno).value

def writeData(file,sheetName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columno).value = data
    workbook.save(file)


def fillGreenColor(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum,columno).fill = greenFill
    workbook.save(file)


def fillRedColor(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum,columno).fill = redFill
    workbook.save(file)



