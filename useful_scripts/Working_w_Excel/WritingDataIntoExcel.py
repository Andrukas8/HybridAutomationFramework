import openpyxl
import os

file = os.getcwd() + "\\testData.xlsx"

# File ---> WorkBook ---> Sheets ---> Cells
workbook = openpyxl.load_workbook(file)
# sheet = workbook["Data"]
sheet = workbook.active # if only one sheet in workbook

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value = "Welcome"


# Writing multiple data:
sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "John"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "Smith"
sheet.cell(2,3).value = "manager"

sheet.cell(3,1).value = 234
sheet.cell(3,2).value = "Dave"
sheet.cell(3,3).value = "QA"

workbook.save(file)
