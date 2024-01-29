import openpyxl
import os

file = os.getcwd() + "\\data.xlsx"

# File ---> WorkBook ---> Sheets ---> Cells
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = 1
cols = 1

while True:
    if sheet.cell(rows,1).value:
        rows += 1
    else:
        rows -= 1
        break

while True:
    if sheet.cell(cols,1).value:
        cols += 1
    else:
        cols -= 1
        break

print(f"Rows = {rows} Cols = {cols}")

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        if sheet.cell(r,c).value:
            print(sheet.cell(r,c).value, end="   ")
        else:
            break
    print()
